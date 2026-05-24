from dotenv import load_dotenv
from pathlib import Path
from contextlib import asynccontextmanager

import io
import os
import uuid
import logging

from datetime import datetime, timezone, timedelta
from typing import Optional, Any

import bcrypt
import jwt
import openpyxl

from motor.motor_asyncio import AsyncIOMotorClient

from fastapi import (
    FastAPI,
    APIRouter,
    HTTPException,
    Depends,
    Request,
)

from fastapi.responses import StreamingResponse

from starlette.middleware.cors import CORSMiddleware

from pydantic import BaseModel, EmailStr, Field

from openpyxl.styles import (
    Font,
    PatternFill,
    Alignment,
    Border,
    Side,
)

# =========================================================
# ENV
# =========================================================

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

MONGO_URL = os.getenv("MONGO_URL")
DB_NAME = os.getenv("DB_NAME", "bayer_db")
JWT_SECRET = os.getenv("JWT_SECRET", "super_secret_key")

if not MONGO_URL:
    raise RuntimeError("MONGO_URL não configurado")

# =========================================================
# DATABASE
# =========================================================

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

# =========================================================
# LOGGING
# =========================================================

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)

logger = logging.getLogger(__name__)

# =========================================================
# CONSTANTS
# =========================================================

JWT_ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7

DEFAULT_PRODUCTS = [
  { "name": "ALSYSTIN", "abbr": "ALS" },
  { "name": "BULLDOCK", "abbr": "BUL" },
  { "name": "CONNECT", "abbr": "CON" },
  { "name": "CURBIX", "abbr": "CUR" },
  { "name": "FOX", "abbr": "FOX" },
  { "name": "FOX XPRO", "abbr": "FXX" },
  { "name": "NATIVO", "abbr": "NAT" },
  { "name": "OBERON", "abbr": "OBE" },
  { "name": "PREMIER PLUS", "abbr": "PRP" },
  { "name": "PROVADO", "abbr": "PRO" },
  { "name": "SPHERE MAX", "abbr": "SPM" },
  { "name": "FINISH", "abbr": "FIN" },
  { "name": "SOBERAN", "abbr": "SOB" },
  { "name": "VERANGO", "abbr": "VER" },
  { "name": "BELT", "abbr": "BEL" },
  { "name": "MOVENTO", "abbr": "MOV" },
  { "name": "DECIS", "abbr": "DEC" }
]

# =========================================================
# HELPERS
# =========================================================


def hash_password(password: str) -> str:
    return bcrypt.hashpw(
        password.encode("utf-8"),
        bcrypt.gensalt(),
    ).decode("utf-8")


def verify_password(
    plain: str,
    hashed: str,
) -> bool:
    return bcrypt.checkpw(
        plain.encode("utf-8"),
        hashed.encode("utf-8"),
    )


def create_access_token(
    user_id: str,
    email: str,
) -> str:
    expire = datetime.now(
        timezone.utc
    ) + timedelta(
        minutes=ACCESS_TOKEN_EXPIRE_MINUTES
    )

    payload = {
        "sub": user_id,
        "email": email,
        "type": "access",
        "exp": int(expire.timestamp()),
    }

    return jwt.encode(
        payload,
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )


def create_reset_token(
    user_id: str,
    email: str,
) -> str:
    expire = datetime.now(
        timezone.utc
    ) + timedelta(
        minutes=30
    )

    payload = {
        "sub": user_id,
        "email": email,
        "type": "reset",
        "exp": int(expire.timestamp()),
    }

    return jwt.encode(
        payload,
        JWT_SECRET,
        algorithm=JWT_ALGORITHM,
    )


def auto_abbreviate(name: str) -> str:
    name = (name or "").strip()

    if not name:
        return ""

    for p in DEFAULT_PRODUCTS:
        if p["name"].lower() == name.lower():
            return p["abbr"]

    cleaned = "".join(
        c for c in name if c.isalpha()
    )

    if cleaned:
        return cleaned[:3].upper()

    return name[:3].upper()


def greeting_for_now() -> str:
    now = datetime.now(
        timezone.utc
    ) - timedelta(hours=3)

    hour = now.hour

    if 5 <= hour < 12:
        return "Bom dia"

    if 12 <= hour < 18:
        return "Boa tarde"

    return "Boa noite"


# =========================================================
# MODELS
# =========================================================


class UserPublic(BaseModel):
    id: str
    email: str
    name: str
    role: str = "user"


class RegisterRequest(BaseModel):
    email: EmailStr
    password: str = Field(min_length=6)
    name: str = Field(min_length=1)


class LoginRequest(BaseModel):
    email: EmailStr
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserPublic


class ForgotPasswordRequest(BaseModel):
    email: EmailStr


class ResetPasswordRequest(BaseModel):
    token: str
    password: str = Field(min_length=6)


class PasswordResetResponse(BaseModel):
    message: str = "Senha redefinida com sucesso"


class ProductionItem(BaseModel):
    id: str = Field(
        default_factory=lambda: str(uuid.uuid4())
    )

    date: str
    unit: str
    sc: str

    product: str
    product_abbr: str = ""

    batch: str

    quantity: Optional[float] = None
    quantity_unit: str = "bag"

    material_status: str = "Disponível"
    situation: str = "A preparar"

    observation: str = ""

    created_at: str = Field(
        default_factory=lambda:
        datetime.now(timezone.utc).isoformat()
    )

    updated_at: str = Field(
        default_factory=lambda:
        datetime.now(timezone.utc).isoformat()
    )


class ProductionItemCreate(BaseModel):
    date: str
    unit: str
    sc: str

    product: str
    batch: str

    quantity: Optional[float] = None
    quantity_unit: str = "kg"

    material_status: str = "Disponível"
    situation: str = "A preparar"

    observation: str = ""


class ProductionItemUpdate(BaseModel):
    unit: Optional[str] = None
    sc: Optional[str] = None

    product: Optional[str] = None
    batch: Optional[str] = None

    quantity: Optional[float] = None
    quantity_unit: Optional[str] = None

    material_status: Optional[str] = None
    situation: Optional[str] = None

    observation: Optional[str] = None
    date: Optional[str] = None


class ProductCreate(BaseModel):
    name: str
    abbr: Optional[str] = None


class ReportRequest(BaseModel):
    date: str
    extra_observations: Optional[str] = None


class Recipe(BaseModel):
    name: str
    abbr: str
    description: Optional[str] = None
    ingredients: Optional[list[str]] = None
    procedure: Optional[str] = None
    category: str = "produtos"


class GalleryImage(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    user_id: str
    url: str
    caption: Optional[str] = None
    category: str = "geral"
    created_at: str = Field(
        default_factory=lambda:
        datetime.now(timezone.utc).isoformat()
    )


# =========================================================
# APP
# =========================================================


@asynccontextmanager
async def lifespan(app: FastAPI):

    await db.users.create_index(
        "email",
        unique=True,
    )

    await db.users.create_index(
        "id",
        unique=True,
    )

    await db.production_items.create_index(
        "id",
        unique=True,
    )

    await db.production_items.create_index(
        "date"
    )

    await db.products.create_index(
        "name",
        unique=True,
    )

    await db.password_resets.create_index(
        "user_id",
        unique=True,
    )

    await db.gallery.create_index(
        "id",
        unique=True,
    )

    await db.gallery.create_index(
        "created_at"
    )

    admin_email = os.getenv(
        "ADMIN_EMAIL",
        "admin@bayer.com",
    ).lower()

    admin_password = os.getenv(
        "ADMIN_PASSWORD",
        "admin123",
    )

    existing = await db.users.find_one({
        "email": admin_email
    })

    if existing is None:

        await db.users.insert_one({
            "id": str(uuid.uuid4()),
            "email": admin_email,
            "password_hash": hash_password(
                admin_password
            ),
            "name": "Administrador",
            "role": "admin",
            "created_at": datetime.now(
                timezone.utc
            ).isoformat(),
        })

        logger.info(
            "Admin criado: %s",
            admin_email,
        )

    yield

    client.close()


app = FastAPI(
    title="Bayer Production Control",
    lifespan=lifespan,
)

api_router = APIRouter(
    prefix="/api"
)

# =========================================================
# CORS
# =========================================================

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.getenv(
        "CORS_ORIGINS",
        "*",
    ).split(","),
    allow_methods=["*"],
    allow_headers=["*"],
)

# =========================================================
# AUTH
# =========================================================


async def get_current_user(
    request: Request,
) -> dict[str, Any]:

    auth_header = request.headers.get(
        "Authorization",
        "",
    )

    token = ""

    if auth_header.startswith("Bearer "):
        token = auth_header[7:]

    if not token:
        raise HTTPException(
            status_code=401,
            detail="Não autenticado",
        )

    try:

        payload = jwt.decode(
            token,
            JWT_SECRET,
            algorithms=[JWT_ALGORITHM],
        )

        if payload.get("type") != "access":
            raise HTTPException(
                status_code=401,
                detail="Token inválido",
            )

    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=401,
            detail="Token expirado",
        )

    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=401,
            detail="Token inválido",
        )

    user = await db.users.find_one(
        {"id": payload["sub"]},
        {"_id": 0, "password_hash": 0},
    )

    if not user:
        raise HTTPException(
            status_code=401,
            detail="Usuário não encontrado",
        )

    return user


# =========================================================
# ROUTES
# =========================================================


@api_router.get("/")
async def root():
    return {
        "service": "Bayer Production Control",
        "ok": True,
    }


@api_router.get("/health")
async def health():
    return {
        "status": "ok",
        "timestamp": datetime.now(
            timezone.utc
        ),
    }


# =========================================================
# AUTH ROUTES
# =========================================================


@api_router.post(
    "/auth/register",
    response_model=TokenResponse,
)
async def register(
    payload: RegisterRequest
):

    email = payload.email.lower()

    existing = await db.users.find_one({
        "email": email
    })

    if existing:
        raise HTTPException(
            status_code=400,
            detail="E-mail já cadastrado",
        )

    user_id = str(uuid.uuid4())

    await db.users.insert_one({
        "id": user_id,
        "email": email,
        "password_hash": hash_password(
            payload.password
        ),
        "name": payload.name,
        "role": "user",
        "created_at": datetime.now(
            timezone.utc
        ).isoformat(),
    })

    token = create_access_token(
        user_id,
        email,
    )

    return TokenResponse(
        access_token=token,
        user=UserPublic(
            id=user_id,
            email=email,
            name=payload.name,
            role="user",
        ),
    )


@api_router.post(
    "/auth/login",
    response_model=TokenResponse,
)
async def login(
    payload: LoginRequest
):

    email = payload.email.lower()

    user = await db.users.find_one({
        "email": email
    })

    if not user:
        raise HTTPException(
            status_code=401,
            detail="E-mail ou senha inválidos",
        )

    if not verify_password(
        payload.password,
        user["password_hash"],
    ):
        raise HTTPException(
            status_code=401,
            detail="E-mail ou senha inválidos",
        )

    token = create_access_token(
        user["id"],
        email,
    )

    return TokenResponse(
        access_token=token,
        user=UserPublic(
            id=user["id"],
            email=email,
            name=user.get("name", ""),
            role=user.get("role", "user"),
        ),
    )


@api_router.get(
    "/auth/me",
    response_model=UserPublic,
)
async def me(
    user: dict = Depends(
        get_current_user
    )
):

    return UserPublic(
        id=user["id"],
        email=user["email"],
        name=user.get("name", ""),
        role=user.get("role", "user"),
    )


@api_router.post("/auth/forgot-password")
async def forgot_password(
    payload: ForgotPasswordRequest,
):

    email = payload.email.lower()

    user = await db.users.find_one({
        "email": email
    })

    if not user:
        return {
            "message": "Se o e-mail existir, um link de reset será enviado"
        }

    reset_token = create_reset_token(
        user["id"],
        email,
    )

    await db.password_resets.update_one(
        {"user_id": user["id"]},
        {
            "$set": {
                "user_id": user["id"],
                "email": email,
                "token": reset_token,
                "created_at": datetime.now(
                    timezone.utc
                ).isoformat(),
                "used": False,
            }
        },
        upsert=True,
    )

    logger.info(
        "Reset token criado para: %s",
        email,
    )

    return {
        "message": "Se o e-mail existir, um link de reset será enviado",
        "reset_token": reset_token,
    }


@api_router.post(
    "/auth/reset-password",
    response_model=PasswordResetResponse,
)
async def reset_password(
    payload: ResetPasswordRequest,
):

    try:

        decoded = jwt.decode(
            payload.token,
            JWT_SECRET,
            algorithms=[JWT_ALGORITHM],
        )

        if decoded.get("type") != "reset":
            raise HTTPException(
                status_code=400,
                detail="Token inválido",
            )

    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=400,
            detail="Link expirado. Solicite um novo reset",
        )

    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=400,
            detail="Token inválido",
        )

    user_id = decoded.get("sub")

    user = await db.users.find_one({
        "id": user_id
    })

    if not user:
        raise HTTPException(
            status_code=404,
            detail="Usuário não encontrado",
        )

    reset_record = await db.password_resets.find_one({
        "user_id": user_id,
    })

    if reset_record and reset_record.get("used"):
        raise HTTPException(
            status_code=400,
            detail="Este link de reset já foi utilizado",
        )

    await db.users.update_one(
        {"id": user_id},
        {
            "$set": {
                "password_hash": hash_password(
                    payload.password
                ),
                "updated_at": datetime.now(
                    timezone.utc
                ).isoformat(),
            }
        },
    )

    await db.password_resets.update_one(
        {"user_id": user_id},
        {
            "$set": {
                "used": True,
                "used_at": datetime.now(
                    timezone.utc
                ).isoformat(),
            }
        },
    )

    logger.info(
        "Senha resetada para: %s",
        user["email"],
    )

    return PasswordResetResponse()


# =========================================================
# ITEMS
# =========================================================


@api_router.get("/items")
async def list_items(
    date: Optional[str] = None,
    user: dict = Depends(get_current_user),
):

    query = {}

    if date:
        query["date"] = date

    items = await db.production_items.find(
        query,
        {"_id": 0},
    ).sort(
        "created_at",
        1,
    ).to_list(2000)

    return items


@api_router.post(
    "/items",
    response_model=ProductionItem,
)
async def create_item(
    payload: ProductionItemCreate,
    user: dict = Depends(get_current_user),
):

    item = ProductionItem(
        **payload.model_dump(),
        product_abbr=auto_abbreviate(
            payload.product
        ),
    )

    await db.production_items.insert_one(
        item.model_dump()
    )

    return item


@api_router.put(
    "/items/{item_id}",
    response_model=ProductionItem,
)
async def update_item(
    item_id: str,
    payload: ProductionItemUpdate,
    user: dict = Depends(get_current_user),
):

    existing = await db.production_items.find_one(
        {"id": item_id},
        {"_id": 0},
    )

    if not existing:
        raise HTTPException(
            status_code=404,
            detail="Item não encontrado",
        )

    update = {
        k: v
        for k, v in payload.model_dump(
            exclude_unset=True
        ).items()
        if v is not None
    }

    if "product" in update:
        update["product_abbr"] = auto_abbreviate(
            update["product"]
        )

    update["updated_at"] = datetime.now(
        timezone.utc
    ).isoformat()

    await db.production_items.update_one(
        {"id": item_id},
        {"$set": update},
    )

    merged = {
        **existing,
        **update,
    }

    return ProductionItem(**merged)


@api_router.delete("/items/{item_id}")
async def delete_item(
    item_id: str,
    user: dict = Depends(get_current_user),
):

    result = await db.production_items.delete_one({
        "id": item_id
    })

    if result.deleted_count == 0:
        raise HTTPException(
            status_code=404,
            detail="Item não encontrado",
        )

    return {"ok": True}


# =========================================================
# PRODUCTS
# =========================================================


@api_router.get("/products")
async def list_products(
    user: dict = Depends(get_current_user),
):

    custom = await db.products.find(
        {},
        {"_id": 0},
    ).to_list(500)

    merged = {
        p["name"]: p["abbr"]
        for p in DEFAULT_PRODUCTS
    }

    for item in custom:
        merged[item["name"]] = item["abbr"]

    return [
        {
            "name": k,
            "abbr": v,
        }
        for k, v in merged.items()
    ]


@api_router.post("/products")
async def add_product(
    payload: ProductCreate,
    user: dict = Depends(get_current_user),
):

    abbr = (
        payload.abbr
        or auto_abbreviate(payload.name)
    ).upper()

    await db.products.update_one(
        {"name": payload.name},
        {
            "$set": {
                "name": payload.name,
                "abbr": abbr,
            }
        },
        upsert=True,
    )

    return {
        "name": payload.name,
        "abbr": abbr,
    }


# =========================================================
# RECIPES
# =========================================================

DEFAULT_RECIPES = [
    {
        "name": "VERANGO",
        "abbr": "VER",
        "category": "produtos",
        "description": "Fungicida sistêmico para folhas",
        "ingredients": ["Trifloxystrobin", "Fluopyram"],
        "procedure": "Aplicar de acordo com recomendações técnicas",
    },
    {
        "name": "NATIVO",
        "abbr": "NAT",
        "category": "produtos",
        "description": "Fungicida de contato e sistêmico",
        "ingredients": ["Trifloxystrobin", "Tebucconazole"],
        "procedure": "Pulverizar uniformemente a cultura",
    },
    {
        "name": "OBERON",
        "abbr": "OBE",
        "category": "produtos",
        "description": "Acaricida seletivo",
        "ingredients": ["Spiromesifen"],
        "procedure": "Indicado para controle de ácaros",
    },
    {
        "name": "FOX XPRO",
        "abbr": "FXX",
        "category": "produtos",
        "description": "Fungicida tríplice ação",
        "ingredients": ["Trifloxystrobin", "Protioconazole", "Bixafen"],
        "procedure": "Aplicação em estágio inicial de infecção",
    },
    {
        "name": "FOX",
        "abbr": "FOX",
        "category": "produtos",
        "description": "Fungicida dupla ação",
        "ingredients": ["Trifloxystrobin", "Protioconazole"],
        "procedure": "Pulverização recomendada",
    },
    {
        "name": "BELT",
        "abbr": "BEL",
        "category": "produtos",
        "description": "Inseticida neonicotinóide",
        "ingredients": ["Flubendiamide"],
        "procedure": "Controle de lepidópteros",
    },
]


@api_router.get("/recipes")
async def list_recipes(
    category: Optional[str] = None,
    user: dict = Depends(get_current_user),
):

    recipes = DEFAULT_RECIPES

    if category:
        recipes = [
            r for r in recipes
            if r.get("category") == category
        ]

    return recipes


# =========================================================
# GALLERY
# =========================================================


@api_router.get("/gallery")
async def list_gallery(
    category: Optional[str] = None,
    skip: int = 0,
    limit: int = 50,
    user: dict = Depends(get_current_user),
):

    query = {}

    if category:
        query["category"] = category

    images = await db.gallery.find(
        query,
        {"_id": 0},
    ).sort(
        "created_at",
        -1,
    ).skip(skip).limit(limit).to_list(limit)

    return images


@api_router.post("/gallery")
async def add_gallery_image(
    payload: dict,
    user: dict = Depends(get_current_user),
):

    image_doc = {
        "id": str(uuid.uuid4()),
        "user_id": user["id"],
        "user_name": user.get("name", ""),
        "url": payload.get("url"),
        "caption": payload.get("caption", ""),
        "category": payload.get("category", "geral"),
        "created_at": datetime.now(
            timezone.utc
        ).isoformat(),
    }

    await db.gallery.insert_one(image_doc)

    return image_doc


@api_router.delete("/gallery/{image_id}")
async def delete_gallery_image(
    image_id: str,
    user: dict = Depends(get_current_user),
):

    img = await db.gallery.find_one({
        "id": image_id
    })

    if not img:
        raise HTTPException(
            status_code=404,
            detail="Imagem não encontrada",
        )

    if img["user_id"] != user["id"] and user.get("role") != "admin":
        raise HTTPException(
            status_code=403,
            detail="Não autorizado",
        )

    await db.gallery.delete_one({
        "id": image_id
    })

    return {"ok": True}


# =========================================================
# REPORT
# =========================================================


def build_report(
    items: list[dict],
    greeting: str,
    extra_obs: Optional[str],
):

    lines = []

    lines.append(
        f"*{greeting}, segue a situação dos materiais para o próximo turno:*"
    )

    lines.append("")

    for item in items:

        qty = ""

        if item.get("quantity") is not None:
            qty = (
                f" - {item['quantity']:g}"
                f"{item.get('quantity_unit', '')}"
            )

        obs = item.get(
            "observation",
            "",
        ).strip()

        obs_text = f" — _{obs}_" if obs else ""

        lines.append(
            f"• {item.get('unit')} | "
            f"{item.get('sc')} | "
            f"{item.get('product')} | "
            f"Lote {item.get('batch')}"
            f"{qty} | "
            f"{item.get('situation')}"
            f"{obs_text}"
        )

    if extra_obs:
        lines.append("")
        lines.append(f"📝 {extra_obs}")

    return "\n".join(lines)


@api_router.post("/reports/whatsapp")
async def whatsapp_report(
    payload: ReportRequest,
    user: dict = Depends(get_current_user),
):

    items = await db.production_items.find(
        {"date": payload.date},
        {"_id": 0},
    ).to_list(2000)

    greeting = greeting_for_now()

    text = build_report(
        items,
        greeting,
        payload.extra_observations,
    )

    return {
        "text": text,
        "count": len(items),
        "greeting": greeting,
    }


# =========================================================
# EXCEL EXPORT
# =========================================================


@api_router.get("/export/excel")
async def export_excel(
    date: str,
    user: dict = Depends(get_current_user),
):

    items = await db.production_items.find(
        {"date": date},
        {"_id": 0},
    ).to_list(2000)

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = f"Bayer {date}"

    headers = [
        "Unidade",
        "SC",
        "Produto",
        "Abrev",
        "Lote",
        "Quantidade",
        "Status MP",
        "Situação",
        "Observação",
    ]

    ws.append(headers)

    header_fill = PatternFill(
        start_color="00A04E",
        end_color="00A04E",
        fill_type="solid",
    )

    bold_white = Font(
        bold=True,
        color="FFFFFF",
    )

    thin = Side(
        border_style="thin",
        color="DDDDDD",
    )

    border = Border(
        left=thin,
        right=thin,
        top=thin,
        bottom=thin,
    )

    for col in range(1, len(headers) + 1):

        cell = ws.cell(
            row=1,
            column=col,
        )

        cell.fill = header_fill
        cell.font = bold_white
        cell.border = border

        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
        )

    for item in items:

        qty = ""

        if item.get("quantity") is not None:
            qty = (
                f"{item['quantity']:g} "
                f"{item.get('quantity_unit', '')}"
            )

        row = [
            item.get("unit", ""),
            item.get("sc", ""),
            item.get("product", ""),
            item.get("product_abbr", ""),
            item.get("batch", ""),
            qty,
            item.get("material_status", ""),
            item.get("situation", ""),
            item.get("observation", ""),
        ]

        ws.append(row)

    widths = [12, 8, 20, 10, 16, 14, 16, 16, 30]

    for i, width in enumerate(widths, 1):
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(i)
        ].width = width

    buffer = io.BytesIO()

    wb.save(buffer)

    buffer.seek(0)

    filename = f"bayer_planilha_{date}.xlsx"

    return StreamingResponse(
        buffer,
        media_type=(
            "application/"
            "vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        headers={
            "Content-Disposition":
            f'attachment; filename="{filename}"'
        },
    )


# =========================================================
# INCLUDE ROUTER
# =========================================================

app.include_router(api_router)
